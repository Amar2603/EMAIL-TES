import re
import smtplib
import dns.resolver
import csv
import io
import os
import threading
from queue import Queue
from time import time, sleep
from datetime import timedelta
from collections import defaultdict
import random
import string
import json
import socket
from concurrent.futures import ThreadPoolExecutor
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs

# Config (high-throughput profile)
NUM_THREADS = 400
OUTPUT_FILE = 'email_validation_results.csv'
NUM_FAKE_CHECKS = 3
SMTP_RETRIES = 5
SMTP_CONNECT_TIMEOUT = 15
MAX_MX_HOSTS = 5
BASE_RETRY_DELAY_SECONDS = 0.25
DOMAIN_CONCURRENCY = 20
DOMAIN_MIN_INTERVAL_SECONDS = 0.0
ENABLE_CATCH_ALL_CHECK = False
MAIL_FROM_CANDIDATES = (
    'postmaster@{domain}',
    'noreply@{domain}',
    'verify@{domain}',
    'bounce@{domain}',
    '<>'
)

# List of common disposable email domains
DISPOSABLE_DOMAINS = {
    '10minutemail.com', 'guerrillamail.com', 'mailinator.com', 'temp-mail.org',
    'throwaway.email', 'yopmail.com', 'maildrop.cc', 'tempail.com', 'dispostable.com',
    'fakeinbox.com', 'getnada.com', 'sharklasers.com', 'spam4.me', 'trashmail.com',
    'mintemail.com', 'mailnesia.com', 'emailondeck.com', 'tempmailo.com', 'tempmail.address'
}

status_counter = defaultdict(int)
lock = threading.Lock()
results = []
start_time = time()
domain_locks = {}
domain_locks_guard = threading.Lock()
domain_last_attempt = defaultdict(float)
mx_cache = {}
mx_cache_lock = threading.Lock()

def get_domain_lock(domain):
    key = (domain or '').lower()
    with domain_locks_guard:
        if key not in domain_locks:
            domain_locks[key] = threading.Semaphore(DOMAIN_CONCURRENCY)
        return domain_locks[key]

def validate_email_format(email):
    # Stricter regex pattern similar to singleemail.py
    pattern = r"^[a-zA-Z0-9]([a-zA-Z0-9._-]*[a-zA-Z0-9])?@[a-zA-Z0-9]([a-zA-Z0-9.-]*[a-zA-Z0-9])?\.[a-zA-Z]{2,}$"
    return re.match(pattern, email) is not None

def extract_emails_from_text(text):
    emails = []
    seen = set()
    for token in re.split(r"[\s,;]+", text or ""):
        email = token.strip().strip("<>()[]{}\"'")
        if email and validate_email_format(email):
            lower_email = email.lower()
            if lower_email not in seen:
                seen.add(lower_email)
                emails.append(email)
    return emails

def extract_emails_from_csv_text(csv_text):
    lines = (csv_text or "").splitlines()
    if not lines:
        return []

    emails = []
    seen = set()

    reader = csv.DictReader(lines)
    if reader.fieldnames:
        field_lookup = {str(name).strip().lower(): name for name in reader.fieldnames if name is not None}
        email_field = field_lookup.get('email')
        for row in reader:
            if email_field:
                candidates = [row.get(email_field, '')]
            else:
                candidates = row.values()

            for value in candidates:
                for email in extract_emails_from_text(str(value or '')):
                    lower_email = email.lower()
                    if lower_email not in seen:
                        seen.add(lower_email)
                        emails.append(email)
    else:
        plain_reader = csv.reader(lines)
        for row in plain_reader:
            for value in row:
                for email in extract_emails_from_text(str(value or '')):
                    lower_email = email.lower()
                    if lower_email not in seen:
                        seen.add(lower_email)
                        emails.append(email)

    return emails

def extract_emails_from_xlsx_bytes(file_bytes):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError('XLSX support requires openpyxl. Run: pip install openpyxl')

    emails = []
    seen = set()

    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    for ws in wb.worksheets:
        header_map = {}
        first_row = True
        for row in ws.iter_rows(values_only=True):
            values = [str(cell).strip() if cell is not None else '' for cell in row]
            if first_row:
                first_row = False
                header_map = {v.lower(): i for i, v in enumerate(values) if v}
                if 'email' in header_map:
                    idx = header_map['email']
                    candidate = values[idx] if idx < len(values) else ''
                    for email in extract_emails_from_text(candidate):
                        lower_email = email.lower()
                        if lower_email not in seen:
                            seen.add(lower_email)
                            emails.append(email)
                    continue

            for value in values:
                for email in extract_emails_from_text(value):
                    lower_email = email.lower()
                    if lower_email not in seen:
                        seen.add(lower_email)
                        emails.append(email)

    return emails

def extract_emails_from_upload(filename, file_bytes):
    ext = os.path.splitext((filename or '').lower())[1]
    if ext in ('.csv', ''):
        return extract_emails_from_csv_text(file_bytes.decode('utf-8-sig', errors='ignore'))
    if ext == '.txt':
        return extract_emails_from_text(file_bytes.decode('utf-8-sig', errors='ignore'))
    if ext == '.xlsx':
        return extract_emails_from_xlsx_bytes(file_bytes)
    raise ValueError('Unsupported file type. Use .csv, .txt, or .xlsx')

def generate_random_email(domain):
    local_part = ''.join(random.choices(string.ascii_lowercase + string.digits, k=10))
    return f"{local_part}@{domain}"

def validate_domain(domain):
    """Validate domain with multiple fallback methods and return record info."""
    domain_key = (domain or '').lower()

    with mx_cache_lock:
        cached = mx_cache.get(domain_key)
    if cached is not None:
        return cached

    errors = []

    try:
        records = dns.resolver.resolve(domain, 'MX', lifetime=5)
        mx_hosts = sorted(
            (
                (int(getattr(record, 'preference', 0)), str(record.exchange).rstrip('.'))
                for record in records
            ),
            key=lambda item: item[0]
        )
        unique_hosts = []
        seen = set()
        for _, host in mx_hosts:
            host_key = host.lower()
            if host_key not in seen:
                seen.add(host_key)
                unique_hosts.append(host)
            if len(unique_hosts) >= MAX_MX_HOSTS:
                break

        if unique_hosts:
            result = {'type': 'MX', 'record': unique_hosts[0], 'all_mx': unique_hosts, 'errors': errors}
            with mx_cache_lock:
                mx_cache[domain_key] = result
            return result
    except dns.resolver.NXDOMAIN:
        errors.append('MX: Domain does not exist (NXDOMAIN)')
    except dns.resolver.NoAnswer:
        errors.append('MX: No MX record found')
    except dns.resolver.Timeout:
        errors.append('MX: DNS query timeout')
    except Exception as e:
        errors.append(f'MX: {str(e)}')

    for record_type in ('A', 'AAAA', 'NS'):
        try:
            records = dns.resolver.resolve(domain, record_type, lifetime=5)
            record_value = str(records[0]).rstrip('.')
            result = {'type': record_type, 'record': record_value, 'all_mx': [], 'errors': errors}
            with mx_cache_lock:
                mx_cache[domain_key] = result
            return result
        except dns.resolver.NXDOMAIN:
            errors.append(f'{record_type}: Domain does not exist (NXDOMAIN)')
        except dns.resolver.NoAnswer:
            errors.append(f'{record_type}: No {record_type} record found')
        except dns.resolver.Timeout:
            errors.append(f'{record_type}: DNS query timeout')
        except Exception as e:
            errors.append(f'{record_type}: {str(e)}')

    try:
        ip_address = socket.gethostbyname(domain)
        result = {'type': 'SOCKET', 'record': ip_address, 'all_mx': [], 'errors': errors}
        with mx_cache_lock:
            mx_cache[domain_key] = result
        return result
    except socket.gaierror as e:
        errors.append(f'SOCKET: {str(e)}')
    except Exception as e:
        errors.append(f'SOCKET: {str(e)}')

    result = {'type': None, 'record': None, 'all_mx': [], 'errors': errors}
    with mx_cache_lock:
        mx_cache[domain_key] = result
    return result

def _set_common_headers(handler, content_type='application/json'):
    handler.send_header('Content-type', content_type)
    handler.send_header('Access-Control-Allow-Origin', '*')
    handler.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
    handler.send_header('Access-Control-Allow-Headers', 'Content-Type')

def _read_multipart_file(headers, body):
    from email.parser import BytesParser
    from email.policy import default

    content_type = headers.get('Content-Type', '')
    if 'multipart/form-data' not in content_type.lower():
        raise ValueError('Content-Type must be multipart/form-data')

    message = BytesParser(policy=default).parsebytes(
        f'Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n'.encode('utf-8') + body
    )

    for part in message.iter_parts():
        if part.get_content_disposition() != 'form-data':
            continue
        if part.get_param('name', header='content-disposition') != 'file':
            continue
        filename = part.get_filename() or ''
        payload = part.get_payload(decode=True) or b''
        return filename, payload

    raise ValueError('No file uploaded')

def normalize_uploaded_result(status, code, message):
    normalized_status = str(status or '').strip()
    normalized_message = str(message or '').strip()
    lower_message = normalized_message.lower()

    if normalized_status == 'Valid' and (
        lower_message.startswith('email format valid, mail server reachable but mailbox could not be confirmed')
    ):
        return 'Unknown', code, normalized_message

    return normalized_status, code, normalized_message

def process_uploaded_emails(filename, file_bytes):
    from singleemail import validate_email as validate_single_email

    emails = extract_emails_from_upload(filename, file_bytes)
    if not emails:
        raise ValueError('No emails found in uploaded file. Ensure it contains email addresses.')

    worker_count = max(1, min(NUM_THREADS, len(emails)))
    ordered_results = [None] * len(emails)

    def validate_one(item):
        index, email = item
        status, code, message = validate_single_email(email)
        status, code, message = normalize_uploaded_result(status, code, message)
        return index, {
            'email': email,
            'status': status,
            'smtp_code': code,
            'message': message,
        }

    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        for index, result in executor.map(validate_one, enumerate(emails)):
            ordered_results[index] = result

    return ordered_results

class RequestHandler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        _set_common_headers(self)
        self.end_headers()

    def do_GET(self):
        parsed_path = urlparse(self.path)

        if parsed_path.path == '/status':
            self.send_response(200)
            _set_common_headers(self)
            self.end_headers()
            self.wfile.write(json.dumps({'status': 'Server running'}).encode('utf-8'))
            return

        self.send_response(404)
        _set_common_headers(self)
        self.end_headers()
        self.wfile.write(json.dumps({'error': 'Not found'}).encode('utf-8'))

    def do_POST(self):
        parsed_path = urlparse(self.path)
        if parsed_path.path != '/upload-csv':
            self.send_response(404)
            _set_common_headers(self)
            self.end_headers()
            self.wfile.write(json.dumps({'error': 'Not found'}).encode('utf-8'))
            return

        try:
            content_length = int(self.headers.get('Content-Length', '0'))
            body = self.rfile.read(content_length)
            filename, file_bytes = _read_multipart_file(self.headers, body)
            results = process_uploaded_emails(filename, file_bytes)
            self.send_response(200)
            _set_common_headers(self)
            self.end_headers()
            self.wfile.write(json.dumps({'results': results, 'total': len(results)}).encode('utf-8'))
        except ValueError as e:
            self.send_response(400)
            _set_common_headers(self)
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e)}).encode('utf-8'))
        except Exception as e:
            self.send_response(500)
            _set_common_headers(self)
            self.end_headers()
            self.wfile.write(json.dumps({'error': f'Failed to process file: {str(e)}'}).encode('utf-8'))

    def log_message(self, format, *args):
        print(f'[{self.log_date_time_string()}] {format % args}')

def run_server():
    port = 8001
    server = HTTPServer(('localhost', port), RequestHandler)
    print('=' * 60)
    print('E-fy File Upload Verification Server')
    print(f'Running at: http://localhost:{port}')
    print('Upload endpoint: POST /upload-csv')
    print('Health check: GET /status')
    print('=' * 60)
    server.serve_forever()

if __name__ == '__main__':
    run_server()
